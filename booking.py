from datetime import datetime, timedelta
import random
import sys
import time
from openpyxl import load_workbook
from playwright.sync_api import expect, sync_playwright

URL = "https://script.google.com/a/macros/banksinarmas.com/s/AKfycbyGVQZaMoU4Q4HOS51V2Tmt_nnO2UNu4QCfUbk6EWuGVYtamrhMMLoUv-kI1oGHU9-0Nw/exec?v=bookWorkSite"


# =====================================================
# RANDOM DELAY FUNCTIONS
# =====================================================


def random_delay(min_sec=2, max_sec=7):
    delay = random.uniform(min_sec, max_sec)
    print(f"⏳ Delay {delay:.2f} detik...")
    time.sleep(delay)


def short_delay(min_ms=500, max_ms=1500):
    time.sleep(random.uniform(min_ms, max_ms) / 1000)


# =====================================================
# LOAD BOOKING DATA
# =====================================================


def load_booking_data():
    try:
        wb = load_workbook("booking.xlsx")
    except FileNotFoundError:
        print("❌ Error: File 'booking.xlsx' tidak ditemukan.")
        sys.exit(1)

    ws = wb.active
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        data.append(
            {
                "NIK": str(row[0]).strip(),
                "NAMA": str(row[1]).strip(),
                "DIVISI": str(row[2]).strip(),
                "EMAIL": str(row[3]).strip(),
                "WORKSITE": str(row[4]).strip(),
            }
        )

    return data


# =====================================================
# GENERATE SENIN - JUMAT MINGGU DEPAN
# =====================================================


def get_next_week_range():
    today = datetime.today().date()
    days_until_next_monday = 7 - today.weekday()

    next_monday = today + timedelta(days=days_until_next_monday)
    next_friday = next_monday + timedelta(days=4)

    start_date_str = next_monday.strftime("%b %d, %Y")
    end_date_str = next_friday.strftime("%b %d, %Y")

    return start_date_str, end_date_str


# =====================================================
# MAIN EXECUTION
# =====================================================

booking_data = load_booking_data()
START_DATE, END_DATE = get_next_week_range()

print("\n📅 Periode Booking Minggu Depan:")
print(f"   Tanggal Mulai (Senin)  : {START_DATE}")
print(f"   Tanggal Selesai (Jumat): {END_DATE}\n")

results = []

# =====================================================
# PLAYWRIGHT AUTOMATION
# =====================================================

with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)

    for idx, user in enumerate(booking_data, start=1):
        print("\n" + "=" * 60)
        print(f"[{idx}/{len(booking_data)}] Processing Booking : {user['NAMA']}")

        page = browser.new_page()

        try:
            page.goto(URL, wait_until="networkidle")
            page.wait_for_timeout(random.randint(2000, 4000))

            # -------------------------------------------------
            # 1. CARI FRAME DI GOOGLE APPS SCRIPT
            # -------------------------------------------------
            frame = None
            for f in page.frames:
                if f.locator("#nik").count() > 0:
                    frame = f
                    break

            if frame is None:
                print("❌ Form booking tidak ditemukan di dalam frame.")
                results.append(
                    {"name": user["NAMA"], "status": "FAILED (Frame Not Found)"}
                )
                page.close()
                continue

            # -------------------------------------------------
            # 2. INPUT DATA UTAMA
            # -------------------------------------------------
            frame.locator("#nik").fill(user["NIK"])
            short_delay()

            frame.locator("#nama").fill(user["NAMA"])
            short_delay()

            frame.locator("#divisi").fill(user["DIVISI"])
            short_delay()

            frame.locator("#email").fill(user["EMAIL"])
            short_delay()

            expect(frame.locator("#nik")).to_have_value(user["NIK"])
            print("✔ Form profil berhasil diisi")

            # -------------------------------------------------
            # 3. PILIH WORKSITE
            # -------------------------------------------------
            frame.evaluate(
                """
                (site)=>{
                    const select = document.querySelector("#workSite");
                    for(const opt of select.options){
                        if(opt.text.trim() == site){
                            select.value = opt.value || opt.text;
                            break;
                        }
                    }
                    select.dispatchEvent(new Event("change", {bubbles:true}));
                    if(window.M){
                        M.FormSelect.init(select);
                    }
                }
                """,
                user["WORKSITE"],
            )

            page.wait_for_timeout(random.randint(1000, 2500))

            # -------------------------------------------------
            # 4. SET TANGGAL (SENIN - JUMAT)
            # -------------------------------------------------
            frame.evaluate(
                """
                (dates)=>{
                    const startInput = document.getElementById("meetingDate");
                    const endInput = document.getElementById("meetingEnd");

                    startInput.value = dates.start;
                    endInput.value = dates.end;

                    startInput.dispatchEvent(new Event("input", {bubbles:true}));
                    endInput.dispatchEvent(new Event("input", {bubbles:true}));

                    startInput.dispatchEvent(new Event("change", {bubbles:true}));
                    endInput.dispatchEvent(new Event("change", {bubbles:true}));

                    if(window.M) {
                        M.updateTextFields();
                    }

                    checkRoom();
                }
                """,
                {"start": START_DATE, "end": END_DATE},
            )

            page.wait_for_timeout(random.randint(3000, 6000))

            # -------------------------------------------------
            # 5. CEK STATUS RUANGAN
            # -------------------------------------------------
            status = frame.locator("#statusRuangan").inner_text()
            print(f"Status Ketersediaan: {status}")

            page.wait_for_timeout(random.randint(4000, 7000))

            # -------------------------------------------------
            # 6. LOGIC SUBMIT + DUAL-DETECTION (DIALOG & DOM)
            # -------------------------------------------------
            if "available" in status.lower():
                print("✔ Ruangan Tersedia (Available)")
                random_delay(2, 4)

                # Event listener non-blocking untuk menangkap alert browser
                dialog_captured = {"message": None}

                def handle_dialog(dialog):
                    print(
                        f"\n🔔 Alert Terdeteksi via Event: '{dialog.message}'"
                    )
                    dialog_captured["message"] = dialog.message.strip()
                    dialog.accept()

                # Pasang listener dialog
                page.on("dialog", handle_dialog)

                try:
                    # Klik Submit
                    frame.locator("#submit-reservation-detail").click(
                        force=True
                    )
                    print(
                        "✔ Submit diklik, memantau respon server (Max 15 detik)..."
                    )

                    actual_msg = None

                    # Polling 15 detik (Mengecek dialog & DOM secara bersamaan)
                    for _ in range(15):
                        page.wait_for_timeout(1000)

                        # Opsi A: Alert browser berhasil ditangkap listener
                        if dialog_captured["message"]:
                            actual_msg = dialog_captured["message"]
                            break

                        # Opsi B: Pengecekan teks langsung di dalam HTML frame
                        frame_html = frame.content().lower()
                        if "succesfully booked" in frame_html:
                            actual_msg = "Location succesfully booked!"
                            break
                        elif "successfully booked" in frame_html:
                            actual_msg = "Location successfully booked!"
                            break
                        elif "already booked" in frame_html:
                            actual_msg = "Location already booked!"
                            break

                    # Lepas listener dialog agar tidak berpengaruh pada iterasi berikutnya
                    page.remove_listener("dialog", handle_dialog)

                    # Ambil screenshot
                    filename = f"booking_{user['NAMA']}.png"
                    page.screenshot(path=filename, full_page=True)

                    # Evaluasi hasil
                    if actual_msg:
                        print(
                            f"\n💬 RESPONS TERDETEKSI: '{actual_msg}' | Screenshot: {filename}"
                        )
                        actual_msg_lower = actual_msg.lower()

                        if (
                            "succesfully booked" in actual_msg_lower
                            or "successfully booked" in actual_msg_lower
                        ):
                            print(
                                f"✔ Booking SUCCESS Verified untuk {user['NAMA']}"
                            )
                            results.append(
                                {"name": user["NAMA"], "status": "SUCCESS"}
                            )
                        else:
                            print(
                                f"✖ Booking Gagal! Respon Server: '{actual_msg}'"
                            )
                            results.append(
                                {
                                    "name": user["NAMA"],
                                    "status": f"FAILED ({actual_msg})",
                                }
                            )
                    else:
                        print(
                            "✖ Timeout! Tidak ada alert maupun respons teks yang terdeteksi di frame."
                        )
                        results.append(
                            {
                                "name": user["NAMA"],
                                "status": "FAILED (No Response/Timeout)",
                            }
                        )

                except Exception as e:
                    print(f"✖ Error saat proses submit: {e}")
                    results.append(
                        {"name": user["NAMA"], "status": f"FAILED ({e})"}
                    )

            else:
                print("✖ Ruangan Tidak Tersedia (Not Available)")
                results.append(
                    {"name": user["NAMA"], "status": "FAILED (Not Available)"}
                )
                page.screenshot(
                    path=f"booking_{user['NAMA']}_UNAVAILABLE.png",
                    full_page=True,
                )

        except Exception as e:
            print(f"✖ Terjadi error tak terduga pada user {user['NAMA']}: {e}")
            results.append(
                {"name": user["NAMA"], "status": f"FAILED (System Error: {e})"}
            )

        finally:
            page.close()

    browser.close()


# =====================================================
# SUMMARY REPORT
# =====================================================

print("\n" + "=" * 60)
print("📊 RINGKASAN HASIL BOOKING")
print("=" * 60)

success_count = 0
failed_count = 0

for r in results:
    print(f"{r['status']:<35} | {r['name']}")

    if r["status"] == "SUCCESS":
        success_count += 1
    else:
        failed_count += 1

print("\n" + "-" * 60)
print(f"Total Eksekusi : {len(results)}")
print(f"Berhasil       : {success_count}")
print(f"Gagal          : {failed_count}")
print("=" * 60)
